"""CSC spreadsheet generator for bulk worker uploads.

Generates Excel spreadsheets in CSC bulk upload format with user-friendly
column headers and clear validation messages.
"""

import logging
from typing import List
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .automation import WorkerInfo

logger = logging.getLogger(__name__)


class CSCSpreadsheetGenerator:
    """Generates CSC bulk upload spreadsheets.
    
    Creates Excel files in the format expected by CSC bulk upload feature.
    Includes user-friendly formatting with clear headers, instructions,
    and data validation to help users fill out the spreadsheet correctly.
    """
    
    # CSC bulk upload template columns with user-friendly descriptions
    COLUMNS = [
        "Full Name",
        "Email Address",
        "Job Title",
        "Start Date (YYYY-MM-DD)",
        "End Date (YYYY-MM-DD)",
        "Manager Email",
        "Work Location",
        "Office Location",
        "Phone Number"
    ]
    
    # Column descriptions for user guidance
    COLUMN_DESCRIPTIONS = {
        "Full Name": "Worker's full legal name (e.g., John Doe)",
        "Email Address": "Worker's email address (must be valid format)",
        "Job Title": "Worker's job title or role (e.g., Software Engineer)",
        "Start Date (YYYY-MM-DD)": "Contract start date in YYYY-MM-DD format (e.g., 2024-04-01)",
        "End Date (YYYY-MM-DD)": "Contract end date in YYYY-MM-DD format (e.g., 2025-04-01)",
        "Manager Email": "Meta manager's email address",
        "Work Location": "Must be: Remote, Onsite, or Hybrid",
        "Office Location": "Required for Onsite/Hybrid (e.g., Menlo Park, New York)",
        "Phone Number": "Optional: Worker's phone number with country code"
    }
    
    def __init__(self):
        """Initialize spreadsheet generator."""
        pass
    
    def generate(self, workers: List[WorkerInfo], output_path: str) -> str:
        """Generate CSC bulk upload spreadsheet.
        
        Creates a user-friendly Excel spreadsheet with:
        - Clear column headers with descriptions
        - Formatted header row for easy reading
        - Auto-adjusted column widths
        - Instructions worksheet
        
        Args:
            workers: List of worker information
            output_path: Path where spreadsheet will be saved
            
        Returns:
            Path to generated spreadsheet
        """
        logger.info(f"Generating CSC spreadsheet for {len(workers)} workers")
        
        wb = Workbook()
        
        # Create instructions sheet first (will be the first tab users see)
        instructions_ws = wb.active
        instructions_ws.title = "Instructions"
        self._add_instructions_sheet(instructions_ws)
        
        # Create data sheet
        ws = wb.create_sheet("Vendor Workers")
        
        # Add header row with styling
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col_idx, column_name in enumerate(self.COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=column_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Add worker data rows
        for row_idx, worker in enumerate(workers, 2):
            ws.cell(row=row_idx, column=1, value=worker.full_name)
            ws.cell(row=row_idx, column=2, value=worker.email)
            ws.cell(row=row_idx, column=3, value=worker.job_title)
            ws.cell(row=row_idx, column=4, value=worker.start_date)
            ws.cell(row=row_idx, column=5, value=worker.end_date)
            ws.cell(row=row_idx, column=6, value=worker.manager_email)
            ws.cell(row=row_idx, column=7, value=worker.work_location)
            ws.cell(row=row_idx, column=8, value=worker.office_location or "")
            ws.cell(row=row_idx, column=9, value=worker.phone or "")
            
            # Add borders to data cells
            for col_idx in range(1, len(self.COLUMNS) + 1):
                ws.cell(row=row_idx, column=col_idx).border = thin_border
        
        # Auto-adjust column widths for readability
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            # Set width with minimum and maximum bounds for readability
            adjusted_width = min(max(max_length + 2, 15), 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze header row for easy scrolling
        ws.freeze_panes = 'A2'
        
        # Save workbook
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))
        
        logger.info(f"Spreadsheet saved to {output_path}")
        return str(output_path)
    
    def _add_instructions_sheet(self, ws):
        """Add instructions worksheet to help users fill out the spreadsheet.
        
        Args:
            ws: Worksheet object to add instructions to
        """
        # Title
        ws['A1'] = "CSC Vendor Worker Bulk Upload - Instructions"
        ws['A1'].font = Font(bold=True, size=14, color="366092")
        
        # Introduction
        ws['A3'] = "This spreadsheet is used to bulk upload vendor workers to CSC (Contractor Services Center)."
        ws['A4'] = "Please fill out the 'Vendor Workers' sheet with worker details following the guidelines below."
        
        # Column instructions
        ws['A6'] = "Column Instructions:"
        ws['A6'].font = Font(bold=True, size=12)
        
        row = 8
        for col_name, description in self.COLUMN_DESCRIPTIONS.items():
            ws[f'A{row}'] = f"• {col_name}:"
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = description
            row += 1
        
        # Important notes
        ws[f'A{row + 2}'] = "Important Notes:"
        ws[f'A{row + 2}'].font = Font(bold=True, size=12, color="FF0000")
        
        notes = [
            "1. Do NOT modify the header row in the 'Vendor Workers' sheet",
            "2. Email addresses must be valid and unique",
            "3. Dates must be in YYYY-MM-DD format (e.g., 2024-04-01)",
            "4. Work Location must be exactly: Remote, Onsite, or Hybrid",
            "5. Office Location is required for Onsite and Hybrid workers",
            "6. Manager Email must be a valid Meta email address",
            "7. Save the file as .xlsx format before uploading to CSC"
        ]
        
        for idx, note in enumerate(notes, row + 4):
            ws[f'A{idx}'] = note
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 60
    
    def validate_spreadsheet(self, spreadsheet_path: str) -> List[str]:
        """Validate CSC spreadsheet format with user-friendly error messages.
        
        Args:
            spreadsheet_path: Path to spreadsheet to validate
            
        Returns:
            List of validation errors (empty if valid). Each error includes
            clear instructions on how to fix the issue.
        """
        errors = []
        
        try:
            from openpyxl import load_workbook
            wb = load_workbook(spreadsheet_path)
            
            # Check if Vendor Workers sheet exists
            if "Vendor Workers" not in wb.sheetnames:
                errors.append(
                    "Missing 'Vendor Workers' sheet. Please ensure the spreadsheet "
                    "contains a sheet named 'Vendor Workers' with worker data."
                )
                return errors
            
            ws = wb["Vendor Workers"]
            
            # Check header row
            header = [cell.value for cell in ws[1]]
            if header != self.COLUMNS:
                errors.append(
                    f"Invalid header row. The header must exactly match: {', '.join(self.COLUMNS)}. "
                    f"Please do not modify the header row. Download a fresh template if needed."
                )
            
            # Check data rows
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                if not any(row):  # Skip empty rows
                    continue
                
                # Check required fields with helpful messages
                if not row[0]:  # Full Name
                    errors.append(
                        f"Row {row_idx}: Full Name is required. "
                        f"Please enter the worker's full legal name."
                    )
                if not row[1]:  # Email
                    errors.append(
                        f"Row {row_idx}: Email Address is required. "
                        f"Please enter a valid email address."
                    )
                elif row[1] and '@' not in str(row[1]):
                    errors.append(
                        f"Row {row_idx}: Email Address '{row[1]}' is invalid. "
                        f"Please enter a valid email format (e.g., worker@company.com)."
                    )
                if not row[2]:  # Job Title
                    errors.append(
                        f"Row {row_idx}: Job Title is required. "
                        f"Please enter the worker's job title or role."
                    )
                if not row[3]:  # Start Date
                    errors.append(
                        f"Row {row_idx}: Start Date is required. "
                        f"Please enter date in YYYY-MM-DD format (e.g., 2024-04-01)."
                    )
                if not row[4]:  # End Date
                    errors.append(
                        f"Row {row_idx}: End Date is required. "
                        f"Please enter date in YYYY-MM-DD format (e.g., 2025-04-01)."
                    )
                if not row[5]:  # Manager Email
                    errors.append(
                        f"Row {row_idx}: Manager Email is required. "
                        f"Please enter the Meta manager's email address."
                    )
                if not row[6]:  # Work Location
                    errors.append(
                        f"Row {row_idx}: Work Location is required. "
                        f"Must be exactly: Remote, Onsite, or Hybrid."
                    )
                elif row[6] and str(row[6]) not in ["Remote", "Onsite", "Hybrid"]:
                    errors.append(
                        f"Row {row_idx}: Work Location '{row[6]}' is invalid. "
                        f"Must be exactly: Remote, Onsite, or Hybrid (case-sensitive)."
                    )
        
        except Exception as e:
            errors.append(
                f"Failed to read spreadsheet: {e}. "
                f"Please ensure the file is a valid .xlsx Excel file."
            )
        
        return errors
